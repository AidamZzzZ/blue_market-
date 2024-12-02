
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.views.generic import DetailView, TemplateView
from django.db.models import Avg
from django.contrib import messages
from django.core.files.base import ContentFile
from django.db.models import Q
from django.core.paginator import Paginator
from django.http import Http404
import requests
import openpyxl
from openpyxl import Workbook
from django.http.response import HttpResponse

from .forms import (
	ProductsForm, 
	CommentsForm, 
	ReplyForm, 
	ReviewRatingForm, 
	TransactionForm,
	ExcelUploadForm
)
from .models import (
	Product, 
	Comment, 
	ReviewRating, 
	Transaction,
)

def home(request):
	# BUSCADOR

	query = request.GET.get('search', '')
	order_by = request.GET.get('order_by', '')
	page = request.GET.get('page', 1)

	products = Product.objects.all()
	
	if query:
		products = products.filter(
			Q(title__icontains=query) |
			Q(description__icontains=query) |
			Q(price__icontains=query)
		)

	if order_by:
		if order_by == 'price_asc':
			products = products.order_by('price')
		elif order_by == 'price_desc':
			products = products.order_by('-price')
		elif order_by == 'rating':
			products = products.order_by('-ratings')
	
	paginator = Paginator(products, 5)
	try:
		products = paginator.page(page)
	except:
		raise Http404



	data = {
		'products': products,
		'entity': products,
		'paginator': paginator,
		'query': query,
		'order_by': order_by,
	}
	return render(request, 'home.html', data)

# AGREGAR PRODUCTOS (CRUD)
@permission_required('products.add_product', raise_exception=True)
@login_required
def add_product(request):
	data = {
		'form': ProductsForm()
	}

	if request.method == 'POST':
		formulario = ProductsForm(data=request.POST, files=request.FILES)
		if formulario.is_valid():
			publication = formulario.save(commit=False)
			publication.user = request.user
			publication.save()
			return redirect(to="list_products")
		else:
			data["form"] = formulario
	return render(request, 'products/create_product.html', data)

# SUBIR LISTA DE PRODUCTOS DESDE HOJA DE EXCEL
@permission_required('products.add_product', raise_exception=True)
def import_products(request):
	if request.method == 'POST':
		form = ExcelUploadForm(request.POST, request.FILES)
		if form.is_valid():
			excel_file = form.cleaned_data['file']
			wb = openpyxl.load_workbook(excel_file)
			sheet = wb.active

			for row in sheet.iter_rows(min_row=2, values_only=True):
				title, description, price, image_data, stock = row
				product = Product(title=title, description=description, price=price, user=request.user, stock=stock)

				response = requests.get(image_data, stream=True)
				if image_data:
					filename = image_data.split('/')[-1]
					product.image.save(filename, ContentFile(response.content), save=False)

				product.save()

			return redirect(to='list_products')
	else:
		form = ExcelUploadForm()

	return render(request, 'products/import_products.html', {'form':form})

# LISTAR PRODUCTOS
@login_required
@permission_required('products.add_product', raise_exception=True)
def list_product(request):
	def test(item):
		return item.name

	user_logged = request.user
	products = user_logged.product_set.all()
	page = request.GET.get('page', 1)



	grupos = user_logged.groups.all() # Listar todos los permisos de todos los grupos 
	permisos = [] 
	for grupo in grupos: 
		permisos.extend(map(test, grupo.permissions.all()))

	try:
		paginator = Paginator(products, 5)
		products = paginator.page(page)
	except:
		raise Http404

	data = {
		'products': products,
		'entity': products,
		'paginator': paginator,
		'permissions': permisos,
	}

	return render(request, 'products/list_product.html', data)

# REPORTE DE MIS PRODUCTOS
class ReportProductExcel(TemplateView):
	def get(self, *args, **kwargs):
		user_logged = self.request.user
		products = user_logged.product_set.all()
		wb = Workbook()
		ws = wb.active
		ws['B1'] = 'MY PRODUCTS'
		
		ws.merge_cells('B1:E1')

		ws['B3'] = 'Title'	
		ws['C3'] = 'Description'
		ws['D3'] = 'Price'
		ws['E3'] = 'Stock'

		cont = 4

		for product in products:
			ws.cell(row=cont, column=2).value = product.title
			ws.cell(row=cont, column=3).value = product.description
			ws.cell(row=cont, column=4).value = product.price
			ws.cell(row=cont, column=5).value = product.stock
			cont+=1

		name_archive = 'ReportMyProducts.xlsx'
		response = HttpResponse(content_type='application/ms-excel')
		content = 'attachment; filename = {0}'.format(name_archive)
		response['Content-Disposition'] = content
		wb.save(response)
		return response

# EDITAR PRODUCTOS
@permission_required('products.change_product', raise_exception=True)
def modify_product(request, id):
	products = get_object_or_404(Product, id=id)

	data = {
		'form': ProductsForm(instance=products)
	}
	
	if request.method == "POST":
		formulario = ProductsForm(data=request.POST, instance=products, files=request.FILES)
		if formulario.is_valid():
			formulario.save()
			return redirect(to="list_products")

	return render(request, 'products/modify_product.html', data)

# ELIMINAR PRODUCTOS
@permission_required('products.delete_product')
def delete_product(request, id):
	products = get_object_or_404(Product, id=id)

	if request.method == 'POST':
		products.delete()
		return redirect(to='list_products')
		
	return render(request, 'products/confirm_delete.html', {'product': products})

# DETALLE DE PRODUCTO
def product_detail(request, pk):
	product = get_object_or_404(Product, id=pk)
	commentForm = CommentsForm()
	replyForm = ReplyForm()
	reviewForm = ReviewRatingForm()

	product_average = product.ratings.aggregate(Avg('rating'))['rating__avg'] or 0
	data = {
		'product': product,
		'form': commentForm,
		'replyform': ReplyForm,
		'reviewForm':reviewForm,
		'product_average': product_average
	}

	return render(request, 'products/detail_product.html', data)

# COMENTARIOS
@login_required
def comment_sent(request, pk):
	product = get_object_or_404(Product, id=pk)

	if request.method == 'POST':
		form = CommentsForm(request.POST)
		if form.is_valid:
			comment = form.save(commit=False)
			comment.author = request.user
			comment.post = product
			comment.save()
	return redirect('detail_product', product.pk)

# RESPONDER A COMENTARIOS
@login_required
def reply_sent(request, pk):
	comment = get_object_or_404(Comment, id=pk)

	if request.method == 'POST':
		form = ReplyForm(request.POST)
		if form.is_valid:
			reply = form.save(commit=False)
			reply.author = request.user
			reply.parent_comment = comment
			reply.save()
	return redirect('detail_product', comment.post.pk)

# CONFIRMAR UNA COMPRA DE PRODUCTO
@login_required
def confirm_product(request, pk):
	product = get_object_or_404(Product, id=pk)

	if product.user == request.user:
		messages.warning(request, 'No puedes comprar tu propio producto')
		return redirect('detail_product', product.pk)
	if product.stock == 0:
		messages.error(request, 'No hay stock de este producto')
		return redirect('detail_product', product.pk)

	if request.method == 'POST':
		form = TransactionForm(request.POST)
		quantity = int(request.POST.get('quantity'))
		if  quantity > product.stock:
			form.add_error('quantity', 'No hay suficiente stock')
			return render(request, 'products/confirm_product.html', {'form':form, 'product':product})

		if form.is_valid():
			transaction = form.save(commit=False)
			transaction.buyer = request.user
			transaction.seller = product.user
			transaction.product = product
			transaction.quantity = quantity
			transaction.save()
			total_cost = product.price * transaction.quantity

		product.stock -= transaction.quantity
		product.save()
		transaction.save()

		return redirect('confirm_product_user_data', transaction.pk)
	else:
		form = TransactionForm()
	return render(request, 'products/confirm_product.html', {'form':form, 'product':product})

# REDIRECCION A LOS DATOS DEL PRODUCTO COMPRADO
def confirm_product_user_data(request, pk):
	transaction = get_object_or_404(Transaction, pk=pk)
	total_cost = transaction.product.price * transaction.quantity

	data = {
		'transaction': transaction,
		'quantity': transaction.quantity,
		'total_cost':total_cost,
	}
	return render(request, 'products/user_confirm.html', data)

# LISTA DE PRODUCTOS COMPRADOS
@login_required
def purchases_list(request):
	user = request.user
	purchases = user.purchase.all()

	page = request.GET.get('page', 1)

	try:
		paginator = Paginator(purchases, 5)
		purchases = paginator.page(page)
	except:
		raise Http404

	data = {
		'purchases': purchases,
		'entity': purchases,
		'paginator': paginator,
	}
	return render(request, 'products/my_buys.html', data)

# LISTA DE PRODUCTOS VENDIDOS
def sales_list(request):
	user = request.user
	seller = user.seller.all()

	page = request.GET.get('page', 1)

	try:
		paginator = Paginator(seller, 5)
		seller = paginator.page(page)
	except:
		raise Http404

	data = {
		'seller': seller,
		'entity': seller,
		'paginator': paginator
	}
	return render(request, 'products/my_sales.html', data)

# AGREGAR RATING Y COMENTARIO A UN PRODUCTO 
@login_required
def review_rating(request, pk):
	product = get_object_or_404(Product, pk=pk)
	reviews = ReviewRating.objects.filter(product=product)
	if request.method == 'POST':
		review_form = ReviewRatingForm(request.POST)
		if review_form.is_valid():
			review = review_form.save(commit=False)
			review.product = product
			review.user = request.user
			review.save()
		return redirect('detail_product', pk=pk)
	data = {
		"form": ReviewRatingForm(),
		"reviews": reviews,
	}
	return render(request, 'products/review_rating.html', data)
